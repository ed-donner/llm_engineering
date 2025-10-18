import os
import time
import pandas as pd
import re
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openai import OpenAI
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Load environment variables
load_dotenv(override=True)
api_key = os.getenv('OPENAI_API_KEY')

# Validate API Key
if not api_key:
    raise ValueError("No API key was found - please check your .env file.")

# Initialize OpenAI client
openai = OpenAI()

# Set up Selenium WebDriver
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--no-sandbox")
chrome_options.add_argument("--disable-dev-shm-usage")

class Website:
    """Scrapes and processes website content using Selenium."""

    def __init__(self, url: str):
        self.url = url
        self.text = "No content extracted."

        service = Service(executable_path="/opt/homebrew/bin/chromedriver")
        driver = webdriver.Chrome(service=service, options=chrome_options)

        try:
            driver.get(url)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            body_element = driver.find_element(By.TAG_NAME, "body")
            self.text = body_element.text.strip() if body_element else "No content extracted."
        except Exception as e:
            print(f"Error fetching website: {e}")
        finally:
            driver.quit()

    def summarized_text(self, max_length=1500):
        return self.text[:max_length] + ("..." if len(self.text) > max_length else "")

def clean_text(text):
    """
    Cleans extracted text by removing markdown-style formatting.
    """
    text = re.sub(r"###*\s*", "", text)
    text = re.sub(r"\*\*(.*?)\*\*", r"\1", text)
    return text.strip()

# Aspect-specific prompts for concise output
aspect_prompts = {
    "Marketing Strategies": "Summarize the core marketing strategies used on this website in in under 30 words. Do not include a title or introduction.",
    "SEO Keywords": "List only the most relevant SEO keywords from this website, separated by commas. Do not include a title or introduction.",
    "User Engagement Tactics": "List key engagement tactics used on this website (e.g., interactive features, user incentives, social proof). Keep responses to 3-5 bullet points. Do not include a title or introduction.",
    "Call-to-Action Phrases": "List only the most common Call-to-Action phrases used on this website, separated by commas. Do not include a title or introduction.",
    "Branding Elements": "Summarize the brand's tone, style, and positioning in under 30 words.  Do not include a title or introduction.",
    "Competitor Comparison": "Briefly describe how this website differentiates itself from competitors in under 30 words.  Do not include a title or introduction.",
    "Product Descriptions": "List the most important features or benefits of the products/services described on this website in under 30 words.  Do not include a title or introduction.",
    "Customer Reviews Sentiment": "Summarize the overall sentiment of customer reviews in oin under 30 words, highlighting common themes.  Do not include a title or introduction.",
    "Social Media Strategy": "List key social media strategies used on this website, separated by commas. Do not include a title or introduction."
}


def summarize(url: str) -> dict:
    """
    Fetches a website, extracts relevant content, and generates a separate summary for each aspect.

    :param url: The website URL to analyze.
    :return: A dictionary containing extracted information.
    """
    website = Website(url)

    if not website.text or website.text == "No content extracted.":
        return {"URL": url, "Error": "Failed to extract content"}

    extracted_data = {"URL": url}

    for aspect, prompt in aspect_prompts.items():
        try:
            formatted_prompt = f"{prompt} \n\nContent:\n{website.summarized_text()}"
            response = openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are an expert at extracting structured information from website content."},
                    {"role": "user", "content": formatted_prompt}
                ]
            )

            extracted_data[aspect] = clean_text(response.choices[0].message.content)

        except Exception as e:
            extracted_data[aspect] = f"Error generating summary: {e}"

    return extracted_data

def save_to_excel(data_list: list, filename="website_analysis.xlsx"):
    """
    Saves extracted information to an Excel file with proper formatting.

    :param data_list: A list of dictionaries containing extracted website details.
    :param filename: The name of the Excel file to save data.
    """
    df = pd.DataFrame(data_list)

    df.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Wrap text for extracted content
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(filename)
    print(f"Data saved to {filename} with improved formatting.")

# 🔹 LIST OF WEBSITES TO PROCESS
websites = [
    "https://www.gymshark.com/",
]

if __name__ == "__main__":
    print("\nProcessing websites...\n")
    extracted_data_list = []

    for site in websites:
        print(f"Extracting data from {site}...")
        extracted_data = summarize(site)
        extracted_data_list.append(extracted_data)

    save_to_excel(extracted_data_list)
    print("\nAll websites processed successfully!")
